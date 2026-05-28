"""Multi-level header flattening for Excel files."""

import os

FLATTEN_VERSION = "v1"



def _flatten_single_sheet(
    file_path: str, sheet_name: str, output_dir: str, file_hash: str
) -> str | None:
    """
    对单个 sheet 展平。返回 CSV 路径，flat 或失败返回 None。
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # 一次计算 fill_map，所有函数共用
    fill_map = fill_merged_cells(ws)

    # Step 1: 读取预览（含合并单元格填充）
    preview = read_excel_preview(ws, max_rows=20, fill_map=fill_map)

    # Step 2: 检测表头结构
    header_info = detect_header_structure(preview)

    # flat 表头不需要展平
    if header_info["header_type"] == "flat":
        wb.close()
        return None

    # Step 3: 展平表头
    skip_rows = header_info["skip_rows"]
    header_rows = header_info["header_rows"]
    header_block = preview[skip_rows:skip_rows + header_rows]
    flattened_headers = flatten_headers(header_block, header_info)

    # Step 4: 读取数据行（含合并单元格填充）
    data_start_row = header_info["data_start_row"]
    data_rows = read_data_rows(ws, data_start_row, fill_map=fill_map)

    wb.close()

    # Step 5: 保存 CSV（L1 缓存）
    l1_key = f"{file_hash}_flat_{FLATTEN_VERSION}_{sheet_name}.csv"
    cache_path = os.path.join(output_dir, l1_key)
    save_flat_csv(flattened_headers, data_rows, cache_path)

    return cache_path


def flatten_excel_headers(file_path: str, output_dir: str) -> dict[str, str] | None:
    """
    主入口：遍历 Excel 所有 sheet，对多级表头 sheet 展平。
    返回 {sheet_name: flat_csv_path}，所有 sheet 都是 flat 时返回 None。
    """
    import openpyxl
    file_hash = compute_file_hash(file_path)

    wb = openpyxl.load_workbook(file_path)
    try:
        results: dict[str, str] = {}
        for sheet_name in wb.sheetnames:
            flat_csv = _flatten_single_sheet(file_path, sheet_name, output_dir, file_hash)
            if flat_csv:
                results[sheet_name] = flat_csv
    finally:
        wb.close()

    return results if results else None

def read_excel_preview(ws, max_rows: int = 20, fill_map: dict | None = None) -> list[list]:
    """读取前 N 行，应用 fill_merged_cells 填充合并单元格。"""
    if fill_map is None:
        fill_map = fill_merged_cells(ws)
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        # 填充合并单元格（values_only 对合并单元格只返回主格值，其他格返回 None）
        filled_row = []
        for j, val in enumerate(row):
            row_num = i + 1  # ws iter_rows 从第1行开始
            col_num = j + 1
            filled_val = fill_map.get((row_num, col_num), val)
            filled_row.append(filled_val)
        rows.append(filled_row)
    return rows

def fill_merged_cells(ws) -> dict[tuple, any]:
    """返回 { (row, col): value } 映射。"""
    fill_map = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = merged_range.start_cell
        value = top_left.value
        for row in ws.iter_rows(
            min_row=merged_range.min_row, max_row=merged_range.max_row,
            min_col=merged_range.min_col, max_col=merged_range.max_col
        ):
            for cell in row:
                fill_map[(cell.row, cell.column)] = value
    return fill_map

def detect_header_structure(preview: list[list]) -> dict:
    """
    基于统计规则检测表头结构。
    规则：
    - skip_rows: 连续全 None 行数
    - header_rows: 从第1个非空行开始，连续的非数据行行数
    - 非数据行判断: str_ratio > 0.3
    - data_start_row = skip_rows + header_rows + 1 (1-indexed，跳到数据行之后)
    - header_type: "flat"(只有1行表头) | "multi"(多行表头)
    """
    if not preview:
        return {"skip_rows": 0, "header_rows": 1, "data_start_row": 1, "header_type": "flat"}

    skip_rows = 0
    for row in preview:
        non_none = sum(1 for v in row if v is not None)
        if non_none == 0:
            skip_rows += 1
        else:
            break

    header_rows = 1
    header_type = "flat"

    for i in range(skip_rows, len(preview)):
        row = preview[i]
        non_none = sum(1 for v in row if v is not None)
        if non_none == 0:
            continue  # 跳过含 None 的行
        str_count = sum(1 for v in row if isinstance(v, str))
        str_ratio = str_count / len(row) if row else 0
        # 数据行特征: 数字为主(str_ratio <= 0.3) 或 全是数字
        if str_ratio > 0.3:
            header_rows = i - skip_rows + 1
            header_type = "multi" if header_rows > 1 else "flat"
        else:
            break  # 遇到数据行，停止

    return {
        "skip_rows": skip_rows,
        "header_rows": header_rows,
        "data_start_row": skip_rows + header_rows + 1,  # 1-indexed，跳到数据行之后
        "header_type": header_type
    }

def _is_data_row(row: list) -> bool:
    """判断是否为数据行：数字为主（str_ratio <= 0.3）"""
    if not row:
        return False
    non_none = sum(1 for v in row if v is not None)
    if non_none == 0:
        return False
    str_count = sum(1 for v in row if isinstance(v, str))
    str_ratio = str_count / len(row)
    return str_ratio <= 0.3

def _is_all_none(row: list) -> bool:
    return all(v is None for v in row)

def sanitize_header_value(val: str) -> str:
    """清理表头值中的特殊字符，替换为下划线。"""
    result = str(val).strip()
    # 中文标点和特殊符号
    result = result.replace("（", "_").replace("）", "_")
    result = result.replace(":", "_").replace("：", "_")
    result = result.replace("、", "_").replace("。", "_")
    result = result.replace(""", "_").replace(""", "_")
    result = result.replace("'", "_").replace("'", "_")
    result = result.replace("《", "_").replace("》", "_")
    result = result.replace("【", "_").replace("】", "_")
    result = result.replace("——", "_").replace("…", "_")
    # 清理连续或首尾的下划线
    while "__" in result:
        result = result.replace("__", "_")
    result = result.strip("_")
    return result if result else ""

def flatten_headers(header_block: list[list], header_info: dict) -> list[str]:
    """展平多级表头为单一列名。"""
    header_rows = header_info["header_rows"]
    num_cols = max(len(row) for row in header_block) if header_block else 0
    flattened = []

    for col_idx in range(num_cols):
        parts = []
        prev_row_values = None
        for row_idx in range(header_rows - 1, -1, -1):
            if row_idx >= len(header_block) or col_idx >= len(header_block[row_idx]):
                continue
            val = header_block[row_idx][col_idx]

            if val is None:
                continue

            if _is_all_none(header_block[row_idx]):
                break

            if _is_data_row(header_block[row_idx]):
                break

            # 行级别重复检测：跳过与上一行完全相同的行
            current_row_values = tuple(header_block[row_idx])
            if prev_row_values is not None and current_row_values == prev_row_values:
                break  # 重复行，停止向上一行查找
            prev_row_values = current_row_values

            s = sanitize_header_value(val)
            if not parts or parts[-1] != s:
                parts.append(s)

        parts.reverse()
        col_name = "_".join(parts) if parts else f"col_{col_idx + 1}"

        if len(col_name) > 200:
            col_name = col_name[:200]

        if col_name in flattened:
            suffix = 2
            while f"{col_name}_{suffix}" in flattened:
                suffix += 1
            col_name = f"{col_name}_{suffix}"

        flattened.append(col_name)

    return flattened

def read_data_rows(ws, data_start_row: int, fill_map: dict | None = None) -> list[list]:
    """
    从 data_start_row（1-indexed）读取所有数据行。
    """
    if fill_map is None:
        fill_map = fill_merged_cells(ws)
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # iter_rows 从第1行开始，i=0 对应 row 1
        if i >= data_start_row - 1:
            filled_row = []
            for j, val in enumerate(row):
                row_num = i + 1
                col_num = j + 1
                filled_val = fill_map.get((row_num, col_num), val)
                filled_row.append(filled_val)
            rows.append(filled_row)
    return rows

def save_flat_csv(headers: list[str], data_rows: list[list], output_path: str) -> None:
    """保存 CSV，utf-8-sig 编码（兼容 Windows Excel）。"""
    import csv
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data_rows)

def compute_file_hash(file_path: str) -> str:
    """计算文件 SHA256 hash（hex 前 12 位）。"""
    import hashlib
    hasher = hashlib.sha256()
    with open(file_path, "rb") as f:
        while chunk := f.read(8192):
            hasher.update(chunk)
    return hasher.hexdigest()[:12]