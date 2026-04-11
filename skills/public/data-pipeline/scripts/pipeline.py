"""
Data Pipeline Script.

Executes a declarative ETL pipeline defined in a JSON spec file.
Uses DuckDB for all data operations: load → clean → transform → merge → export.
"""

import argparse
import html
import json
import os
import re
import sys
import time

try:
    import duckdb
except ImportError:
    print("ERROR: duckdb is required. Install with: pip install duckdb")
    sys.exit(1)

try:
    import openpyxl  # noqa: F401
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Security helpers
# ---------------------------------------------------------------------------

# Allowed DuckDB types for cast_types action
ALLOWED_SQL_TYPES = frozenset({
    "INTEGER", "INT", "BIGINT", "SMALLINT", "TINYINT",
    "DOUBLE", "FLOAT", "REAL", "DECIMAL", "NUMERIC",
    "VARCHAR", "TEXT", "STRING", "CHAR", "BPCHAR",
    "DATE", "TIMESTAMP", "TIMESTAMPTZ", "TIME", "TIMETZ",
    "BOOLEAN", "BOOL", "BLOB", "BYTEA",
    "INTERVAL", "UUID", "JSON", "HUGEINT", "UHUGEINT",
    "UBIGINT", "USMALLINT", "UTINYINT", "UMEDIUMINT",
})

# Path prefixes that are allowed for file operations
_ALLOWED_PATH_PREFIXES = (
    "/mnt/user-data/",
    "/mnt/skills/",
)


def validate_identifier(name: str, context: str = "identifier") -> str:
    """Validate that a name is a safe SQL identifier (table/column name).

    Only allows alphanumeric characters and underscores.
    Raises ValueError if the name contains unsafe characters.
    """
    if not re.match(r"^[a-zA-Z0-9_]+$", name):
        raise ValueError(f"Invalid {context}: '{name}'. Only alphanumeric characters and underscores are allowed.")
    return name


def validate_path(file_path: str, context: str = "file path") -> str:
    """Validate that a file path stays within allowed directories.

    Raises ValueError if the resolved path escapes allowed prefixes.
    """
    resolved = os.path.realpath(file_path)
    if not any(resolved.startswith(prefix) for prefix in _ALLOWED_PATH_PREFIXES):
        # Also allow paths under CWD (for spec files run locally)
        cwd = os.path.realpath(os.getcwd())
        if not resolved.startswith(cwd + os.sep) and resolved != cwd:
            raise ValueError(
                f"Invalid {context}: '{file_path}' resolves to '{resolved}', "
                f"which is outside allowed directories."
            )
    return file_path


def validate_sql_type(type_name: str) -> str:
    """Validate that a SQL type name is in the allowed whitelist."""
    upper = type_name.upper().strip()
    # Handle DECIMAL(p,s) etc.
    base_type = upper.split("(")[0].strip()
    if base_type not in ALLOWED_SQL_TYPES:
        raise ValueError(
            f"Invalid SQL type: '{type_name}'. Allowed types: {sorted(ALLOWED_SQL_TYPES)}"
        )
    return upper


# ---------------------------------------------------------------------------
# Pipeline action handlers
# ---------------------------------------------------------------------------


def action_load_csv(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Load a CSV file into a DuckDB table."""
    file_path = validate_path(step["file"], "load_csv file")
    table_name = validate_identifier(step["as"], "table name")
    con.execute(f"CREATE TABLE \"{table_name}\" AS SELECT * FROM read_csv_auto(?)", [file_path])
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    return f"load_csv → {table_name} ({cnt:,} rows)"


def action_load_excel(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Load an Excel file (all sheets) into DuckDB tables."""
    import openpyxl

    file_path = validate_path(step["file"], "load_excel file")
    base_name = validate_identifier(step["as"], "table base name")
    con.execute("INSTALL spatial; LOAD spatial;")

    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = wb.sheetnames
    wb.close()

    loaded = []
    for sheet in sheets:
        safe_name = f"{base_name}_{re.sub(r'[^a-zA-Z0-9_]', '_', sheet)}"
        con.execute(
            f"CREATE TABLE \"{safe_name}\" AS SELECT * FROM st_read(?, layer = ?, open_options = ['HEADERS=FORCE', 'FIELD_TYPES=AUTO'])",
            [file_path, sheet],
        )
        cnt = con.execute(f'SELECT COUNT(*) FROM "{safe_name}"').fetchone()[0]
        loaded.append(f"{safe_name} ({cnt:,} rows)")

    return f"load_excel → {', '.join(loaded)}"


def action_load_json(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Load a JSON file (array of objects) into a DuckDB table."""
    file_path = validate_path(step["file"], "load_json file")
    table_name = validate_identifier(step["as"], "table name")
    con.execute(f"CREATE TABLE \"{table_name}\" AS SELECT * FROM read_json_auto(?, format='array_of_objects')", [file_path])
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    return f"load_json → {table_name} ({cnt:,} rows)"


def action_drop_nulls(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Drop rows with null values in specified columns."""
    table = validate_identifier(step["table"], "table name")
    columns = step["columns"]
    for c in columns:
        validate_identifier(c, "column name")
    before = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]

    null_checks = " AND ".join(f'"{c}" IS NOT NULL' for c in columns)
    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT * FROM "{table}" WHERE {null_checks}')
    after = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    diff = before - after
    return f"drop_nulls on {table} ({before:,} → {after:,} rows, -{diff:,})"


def action_fill_nulls(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Fill null values with specified defaults."""
    table = validate_identifier(step["table"], "table name")
    columns: dict[str, str] = step["columns"]
    for c in columns:
        validate_identifier(c, "column name in fill_nulls")

    # Count nulls before
    null_counts: dict[str, int] = {}
    for col in columns:
        cnt = con.execute(f'SELECT COUNT(*) FROM "{table}" WHERE "{col}" IS NULL').fetchone()[0]
        null_counts[col] = cnt

    # Build COALESCE expressions using parameterized defaults
    coalesce_parts = []
    all_cols = con.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = ?", [table]).fetchall()
    # We need to build the COALESCE with literal values since DuckDB doesn't
    # support parameterized defaults inside a SELECT expression list easily.
    # Use the connection's escape mechanism for string defaults.
    for (col_name,) in all_cols:
        if col_name in columns:
            default = columns[col_name]
            # For numeric defaults, use as-is after validation; for strings, use quoted literal
            if isinstance(default, (int, float)):
                coalesce_parts.append(f'COALESCE("{col_name}", {default}) AS "{col_name}"')
            else:
                # Escape single quotes in string defaults to prevent SQL injection
                safe_default = str(default).replace("'", "''")
                coalesce_parts.append(f"COALESCE(\"{col_name}\", '{safe_default}') AS \"{col_name}\"")
        else:
            coalesce_parts.append(f'"{col_name}"')

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT {", ".join(coalesce_parts)} FROM "{table}"')

    filled_info = ", ".join(f"{c}: {null_counts[c]:,} filled" for c in columns)
    after = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"fill_nulls on {table} ({after:,} rows, {filled_info})"


def action_drop_duplicates(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Remove duplicate rows based on specified columns."""
    table = validate_identifier(step["table"], "table name")
    columns = step["columns"]
    for c in columns:
        validate_identifier(c, "column name")
    before = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]

    partition_cols = ", ".join(f'"{c}"' for c in columns)
    all_cols = con.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = ?", [table]).fetchall()
    select_cols = ", ".join(f'"{c[0]}"' for c in all_cols)

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT {select_cols} FROM (SELECT *, ROW_NUMBER() OVER (PARTITION BY {partition_cols}) as _rn FROM "{table}") sub WHERE _rn = 1')
    after = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    diff = before - after
    return f"drop_duplicates on {table} ({before:,} → {after:,} rows, -{diff:,})"


def action_trim_strings(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Trim whitespace from string columns."""
    table = validate_identifier(step["table"], "table name")
    columns = step.get("columns")

    all_cols = con.execute(f"SELECT column_name, data_type FROM information_schema.columns WHERE table_name = ?", [table]).fetchall()

    if not columns:
        columns = [c[0] for c in all_cols if "VARCHAR" in c[1].upper() or "TEXT" in c[1].upper()]

    if not columns:
        cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
        return f"trim_strings on {table} ({cnt:,} rows, no string columns found)"

    select_parts = []
    for col_name, _ in all_cols:
        if col_name in columns:
            select_parts.append(f'TRIM("{col_name}") AS "{col_name}"')
        else:
            select_parts.append(f'"{col_name}"')

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT {", ".join(select_parts)} FROM "{table}"')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"trim_strings on {table} ({cnt:,} rows, {len(columns)} columns trimmed)"


def action_normalize_dates(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Parse date columns into a standard format."""
    table = validate_identifier(step["table"], "table name")
    columns = step["columns"]
    for c in columns:
        validate_identifier(c, "column name")
    step.get("format", "%Y-%m-%d")  # date format hint

    all_cols = con.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = ?", [table]).fetchall()

    select_parts = []
    for (col_name,) in all_cols:
        if col_name in columns:
            select_parts.append(f'TRY_CAST("{col_name}" AS DATE) AS "{col_name}"')
        else:
            select_parts.append(f'"{col_name}"')

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT {", ".join(select_parts)} FROM "{table}"')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"normalize_dates on {table} ({cnt:,} rows, {len(columns)} columns)"


def action_cast_types(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Cast columns to specified data types."""
    table = validate_identifier(step["table"], "table name")
    columns: dict[str, str] = step["columns"]
    for c in columns:
        validate_identifier(c, "column name")
    for t in columns.values():
        validate_sql_type(t)

    all_cols = con.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = ?", [table]).fetchall()

    select_parts = []
    for (col_name,) in all_cols:
        if col_name in columns:
            target_type = columns[col_name]
            select_parts.append(f'TRY_CAST("{col_name}" AS {target_type}) AS "{col_name}"')
        else:
            select_parts.append(f'"{col_name}"')

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT {", ".join(select_parts)} FROM "{table}"')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"cast_types on {table} ({cnt:,} rows, {len(columns)} columns cast)"


def action_rename_columns(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Rename columns in a table."""
    table = validate_identifier(step["table"], "table name")
    mapping: dict[str, str] = step["mapping"]
    for old, new in mapping.items():
        validate_identifier(old, "old column name")
        validate_identifier(new, "new column name")

    all_cols = con.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = ?", [table]).fetchall()

    select_parts = []
    for (col_name,) in all_cols:
        new_name = mapping.get(col_name, col_name)
        if new_name != col_name:
            select_parts.append(f'"{col_name}" AS "{new_name}"')
        else:
            select_parts.append(f'"{col_name}"')

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT {", ".join(select_parts)} FROM "{table}"')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"rename_columns on {table} ({cnt:,} rows, {len(mapping)} columns renamed)"


def action_add_computed_column(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Add a new column computed from a SQL expression.

    NOTE: The 'expr' parameter is intentionally a raw DuckDB SQL expression,
    as documented in SKILL.md ("any valid DuckDB SQL expression"). This is
    by design — the expression is written by the Agent based on user instructions
    and runs inside a sandboxed DuckDB instance.
    """
    table = validate_identifier(step["table"], "table name")
    column = validate_identifier(step["column"], "computed column name")
    expr = step["expr"]

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT *, ({expr}) AS "{column}" FROM "{table}"')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"add_computed_column '{column}' on {table} ({cnt:,} rows)"


def action_group_by(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Group and aggregate a table."""
    table = validate_identifier(step["table"], "table name")
    by_cols = step["by"]
    for c in by_cols:
        validate_identifier(c, "group by column")
    aggregations: dict[str, str] = step["aggregations"]
    for alias in aggregations:
        validate_identifier(alias, "aggregation alias")
    output_table = step.get("as", table)
    if output_table != table:
        validate_identifier(output_table, "output table name")

    by_parts = ", ".join(f'"{c}"' for c in by_cols)
    agg_parts = [f'{expr} AS "{alias}"' for alias, expr in aggregations.items()]
    select = f"{by_parts}, {', '.join(agg_parts)}"

    con.execute(f'CREATE OR REPLACE TABLE "{output_table}" AS SELECT {select} FROM "{table}" GROUP BY {by_parts}')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{output_table}"').fetchone()[0]
    return f"group_by on {table} → {output_table} ({cnt:,} rows)"


def action_join(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Join two tables."""
    left = validate_identifier(step["left"], "left table")
    right = validate_identifier(step["right"], "right table")
    on_cols = step["on"]
    for c in on_cols:
        validate_identifier(c, "join column")
    how = step.get("how", "inner")
    output_table = validate_identifier(step["as"], "output table name")

    on_clause = " AND ".join(f'"{left}"."{c}" = "{right}"."{c}"' for c in on_cols)

    left_cols = con.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = ?", [left]).fetchall()
    right_cols = con.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = ?", [right]).fetchall()

    left_select = ", ".join(f'"{left}"."{c[0]}" AS "{c[0]}"' for c in left_cols)
    right_select = ", ".join(f'"{right}"."{c[0]}" AS "{c[0]}_right"' for c in right_cols if c[0] not in [lc[0] for lc in left_cols])
    all_select = f"{left_select}, {right_select}" if right_select else left_select

    con.execute(f'CREATE TABLE "{output_table}" AS SELECT {all_select} FROM "{left}" {how.upper()} JOIN "{right}" ON {on_clause}')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{output_table}"').fetchone()[0]
    return f"join {left} + {right} → {output_table} ({cnt:,} rows, {how})"


def action_union(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Union (stack) multiple tables."""
    tables = step["tables"]
    for t in tables:
        validate_identifier(t, "union table")
    output_table = validate_identifier(step["as"], "output table name")

    union_parts = [f'SELECT * FROM "{t}"' for t in tables]
    con.execute(f'CREATE TABLE "{output_table}" AS ' + " UNION ALL ".join(union_parts))
    cnt = con.execute(f'SELECT COUNT(*) FROM "{output_table}"').fetchone()[0]
    return f"union {', '.join(tables)} → {output_table} ({cnt:,} rows)"


def action_filter(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Filter rows by a SQL WHERE condition.

    NOTE: The 'condition' parameter is intentionally a raw DuckDB SQL WHERE clause,
    as documented in SKILL.md. This is by design — the condition is written by the
    Agent based on user instructions and runs inside a sandboxed DuckDB instance.
    """
    table = validate_identifier(step["table"], "table name")
    condition = step["condition"]
    before = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]

    con.execute(f'CREATE OR REPLACE TABLE "{table}" AS SELECT * FROM "{table}" WHERE {condition}')
    after = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    diff = before - after
    return f"filter on {table} ({before:,} → {after:,} rows, -{diff:,})"


def action_sample(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Take a random sample of rows."""
    table = validate_identifier(step["table"], "table name")
    output_table = step.get("as", table)
    if output_table != table:
        validate_identifier(output_table, "output table name")
    n = step.get("n")
    fraction = step.get("fraction")

    if n:
        con.execute(f'CREATE OR REPLACE TABLE "{output_table}" AS SELECT * FROM "{table}" USING SAMPLE {int(n)} ROWS')
    elif fraction:
        pct = float(fraction) * 100
        con.execute(f'CREATE OR REPLACE TABLE "{output_table}" AS SELECT * FROM "{table}" USING SAMPLE {pct}%')

    cnt = con.execute(f'SELECT COUNT(*) FROM "{output_table}"').fetchone()[0]
    return f"sample on {table} → {output_table} ({cnt:,} rows)"


def action_top_n(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Get top N rows ordered by a column."""
    table = validate_identifier(step["table"], "table name")
    order_by = validate_identifier(step["order_by"], "order by column")
    n = step.get("n", 10)
    direction = step.get("direction", "DESC")
    output_table = step.get("as", table)
    if output_table != table:
        validate_identifier(output_table, "output table name")

    con.execute(f'CREATE OR REPLACE TABLE "{output_table}" AS SELECT * FROM "{table}" ORDER BY "{order_by}" {direction} LIMIT {int(n)}')
    cnt = con.execute(f'SELECT COUNT(*) FROM "{output_table}"').fetchone()[0]
    return f"top_n on {table} → {output_table} ({cnt:,} rows)"


def action_export_csv(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Export a table to CSV."""
    table = validate_identifier(step["table"], "table name")
    file_path = validate_path(step["file"], "export file path")
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    con.execute(f"COPY \"{table}\" TO ? (HEADER, DELIMITER ',')", [file_path])
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"export_csv → {file_path} ({cnt:,} rows)"


def action_export_json(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Export a table to JSON."""
    table = validate_identifier(step["table"], "table name")
    file_path = validate_path(step["file"], "export file path")
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
    con.execute(f"COPY \"{table}\" TO ? (FORMAT JSON, ARRAY true)", [file_path])
    cnt = con.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    return f"export_json → {file_path} ({cnt:,} rows)"


def action_export_excel(con: duckdb.DuckDBPyConnection, step: dict) -> str:
    """Export a table to Excel via openpyxl."""
    table = validate_identifier(step["table"], "table name")
    file_path = validate_path(step["file"], "export file path")
    os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)

    import openpyxl

    result = con.execute(f'SELECT * FROM "{table}"')
    columns = [desc[0] for desc in result.description]
    rows = result.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table[:31]
    ws.append(columns)
    for row in rows:
        ws.append(row)
    wb.save(file_path)

    return f"export_excel → {file_path} ({len(rows):,} rows)"


# ---------------------------------------------------------------------------
# Action dispatch
# ---------------------------------------------------------------------------

ACTION_MAP: dict[str, callable] = {
    "load_csv": action_load_csv,
    "load_excel": action_load_excel,
    "load_json": action_load_json,
    "drop_nulls": action_drop_nulls,
    "fill_nulls": action_fill_nulls,
    "drop_duplicates": action_drop_duplicates,
    "trim_strings": action_trim_strings,
    "normalize_dates": action_normalize_dates,
    "cast_types": action_cast_types,
    "rename_columns": action_rename_columns,
    "add_computed_column": action_add_computed_column,
    "group_by": action_group_by,
    "join": action_join,
    "union": action_union,
    "filter": action_filter,
    "sample": action_sample,
    "top_n": action_top_n,
    "export_csv": action_export_csv,
    "export_json": action_export_json,
    "export_excel": action_export_excel,
}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Execute a data pipeline defined in a JSON spec file")
    parser.add_argument(
        "--spec-file",
        required=True,
        help="Path to the pipeline spec JSON file",
    )
    args = parser.parse_args()

    spec_path = validate_path(args.spec_file, "spec file")
    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)

    steps = spec.get("steps", [])
    if not steps:
        print("No steps defined in pipeline spec.")
        sys.exit(1)

    con = duckdb.connect()
    report_lines: list[str] = []
    report_lines.append("Pipeline Execution Report")
    report_lines.append("=" * 50)

    start_time = time.time()
    export_count = 0

    for i, step in enumerate(steps, 1):
        action = step.get("action", "")
        handler = ACTION_MAP.get(action)

        if not handler:
            msg = f"Step {i}: UNKNOWN action '{action}' — skipping"
            report_lines.append(msg)
            print(f"WARNING: {msg}")
            continue

        try:
            msg = handler(con, step)
            report_lines.append(f"Step {i}: {msg}")
            print(f"Step {i}: {msg}")
            if action.startswith("export_"):
                export_count += 1
        except Exception as e:
            msg = f"Step {i}: FAILED — {e}"
            report_lines.append(msg)
            print(f"ERROR: {msg}")
            con.close()
            sys.exit(1)

    elapsed = time.time() - start_time
    report_lines.append("=" * 50)
    report_lines.append(f"Done in {elapsed:.1f}s ({export_count} file(s) exported)")

    con.close()
    print("\n" + "\n".join(report_lines))


if __name__ == "__main__":
    main()
