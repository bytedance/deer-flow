"""
Report Template Generator Script.

Reads a template definition (JSON), loads data via DuckDB, and generates
a self-contained HTML report with KPI cards, charts (ECharts), tables,
and text sections.
"""

import argparse
import html
import json
import os
import re
import sys

try:
    import duckdb
except ImportError:
    print("ERROR: duckdb is required. Install with: pip install duckdb")
    sys.exit(1)

try:
    import openpyxl  # noqa: F401
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Security helpers
# ---------------------------------------------------------------------------

_ALLOWED_PATH_PREFIXES = (
    "/mnt/user-data/",
    "/mnt/skills/",
)


def validate_path(file_path: str, context: str = "file path") -> str:
    """Validate that a file path stays within allowed directories."""
    resolved = os.path.realpath(file_path)
    if not any(resolved.startswith(prefix) for prefix in _ALLOWED_PATH_PREFIXES):
        cwd = os.path.realpath(os.getcwd())
        if not resolved.startswith(cwd + os.sep) and resolved != cwd:
            raise ValueError(
                f"Invalid {context}: '{file_path}' resolves to '{resolved}', "
                f"which is outside allowed directories."
            )
    return file_path


def sanitize_identifier(name: str) -> str:
    """Sanitize a string to be a safe SQL identifier."""
    return re.sub(r"[^a-zA-Z0-9_]", "_", name)


def esc(value) -> str:
    """HTML-escape a value for safe embedding in HTML output."""
    return html.escape(str(value), quote=True)


def _safe_float(v, default: float = 0.0) -> float:
    """Safely convert a value to float, returning *default* on failure."""
    try:
        return float(v) if v is not None else default
    except (ValueError, TypeError):
        return default


def _inline_bold(text: str) -> str:
    """Replace ``**bold**`` patterns with ``<strong>bold</strong>``."""
    return re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", text)


# ---------------------------------------------------------------------------
# KPI computation
# ---------------------------------------------------------------------------


def _compute_kpis(con: duckdb.DuckDBPyConnection, table_name: str = "data") -> dict[str, float | int | None]:
    """Dynamically compute KPI values based on the columns that actually exist."""
    result: dict[str, float | int | None] = {}
    try:
        cols = {row[0] for row in con.execute(f"DESCRIBE {table_name}").fetchall()}
    except Exception:
        return result

    if "revenue" in cols:
        try:
            row = con.execute(f"SELECT SUM(revenue) as v FROM {table_name}").fetchone()
            result["kpi_revenue"] = row[0] if row else None
        except Exception:
            result["kpi_revenue"] = None
        try:
            row = con.execute(f"SELECT AVG(revenue) as v FROM {table_name}").fetchone()
            result["kpi_avg"] = row[0] if row else None
        except Exception:
            result["kpi_avg"] = None

    if "customer_id" in cols:
        try:
            row = con.execute(f"SELECT COUNT(DISTINCT customer_id) as v FROM {table_name}").fetchone()
            result["kpi_customers"] = row[0] if row else None
        except Exception:
            result["kpi_customers"] = None

    # Total row count is always available
    try:
        row = con.execute(f"SELECT COUNT(*) as v FROM {table_name}").fetchone()
        result["kpi_orders"] = row[0] if row else None
    except Exception:
        result["kpi_orders"] = None

    # WoW and satisfaction cannot be inferred from generic data
    result["kpi_wow"] = None
    result["kpi_satisfaction"] = None

    return result


# ---------------------------------------------------------------------------
# Theme definitions
# ---------------------------------------------------------------------------

THEMES: dict[str, dict[str, str]] = {
    "light": {
        "bg": "#ffffff",
        "card_bg": "#ffffff",
        "text": "#333333",
        "text_secondary": "#666666",
        "border": "#e0e0e0",
        "kpi_bg": "#f8f9fa",
        "header_bg": "#ffffff",
        "shadow": "0 2px 8px rgba(0,0,0,0.08)",
        "accent": "#2c3e50",
        "palette": '["#5470c6","#91cc75","#fac858","#ee6666","#73c0de","#3ba272"]',
    },
    "dark": {
        "bg": "#1a1a2e",
        "card_bg": "#16213e",
        "text": "#e0e0e0",
        "text_secondary": "#a0a0a0",
        "border": "#2a2a4a",
        "kpi_bg": "#0f3460",
        "header_bg": "#16213e",
        "shadow": "0 2px 8px rgba(0,0,0,0.3)",
        "accent": "#e94560",
        "palette": '["#5470c6","#91cc75","#fac858","#ee6666","#73c0de","#3ba272"]',
    },
    "corporate": {
        "bg": "#f0f4f8",
        "card_bg": "#ffffff",
        "text": "#2c3e50",
        "text_secondary": "#5a6c7d",
        "border": "#d0d9e1",
        "kpi_bg": "#e8eef4",
        "header_bg": "#2c3e50",
        "shadow": "0 2px 6px rgba(44,62,80,0.1)",
        "accent": "#2980b9",
        "palette": '["#2980b9","#27ae60","#f39c12","#e74c3c","#8e44ad","#1abc9c"]',
    },
}

ECHARTS_CDN = "https://cdn.jsdelivr.net/npm/echarts@5.5.0/dist/echarts.min.js"


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_data_files(
    con: duckdb.DuckDBPyConnection,
    data_files: list[str],
) -> None:
    """Load all data files into a single 'data' view (or multiple views)."""
    if not data_files:
        return

    # Load first file as the main 'data' table
    first = validate_path(data_files[0], "data file")
    ext = os.path.splitext(first)[1].lower()
    if ext == ".csv":
        con.execute("CREATE TABLE data AS SELECT * FROM read_csv_auto(?)", [first])
    elif ext in (".xlsx", ".xls"):
        con.execute("INSTALL spatial; LOAD spatial;")
        import openpyxl

        wb = openpyxl.load_workbook(first, read_only=True)
        first_sheet = wb.sheetnames[0]
        wb.close()
        con.execute(
            "CREATE TABLE data AS SELECT * FROM st_read(?, layer = ?, open_options = ['HEADERS=FORCE', 'FIELD_TYPES=AUTO'])",
            [first, first_sheet],
        )
    elif ext == ".json":
        con.execute("CREATE TABLE data AS SELECT * FROM read_json_auto(?, format='array_of_objects')", [first])

    # Load additional files as separate tables
    for f in data_files[1:]:
        validate_path(f, "data file")
        base = sanitize_identifier(os.path.splitext(os.path.basename(f))[0])
        ext = os.path.splitext(f)[1].lower()
        if ext == ".csv":
            con.execute(f"CREATE TABLE \"{base}\" AS SELECT * FROM read_csv_auto(?)", [f])
        elif ext in (".xlsx", ".xls"):
            import openpyxl

            wb = openpyxl.load_workbook(f, read_only=True)
            first_sheet = wb.sheetnames[0]
            wb.close()
            con.execute(
                f"CREATE TABLE \"{base}\" AS SELECT * FROM st_read(?, layer = ?, open_options = ['HEADERS=FORCE', 'FIELD_TYPES=AUTO'])",
                [f, first_sheet],
            )
        elif ext == ".json":
            con.execute(f"CREATE TABLE \"{base}\" AS SELECT * FROM read_json_auto(?, format='array_of_objects')", [f])


# ---------------------------------------------------------------------------
# Section rendering
# ---------------------------------------------------------------------------


def format_value(value, fmt: str) -> str:
    """Format a value for display."""
    if value is None:
        return "N/A"
    if fmt == "currency":
        return f"${value:,.2f}"
    if fmt == "percent":
        return f"{value:.1f}%"
    if isinstance(value, float):
        return f"{value:,.1f}"
    return f"{value:,}"


def render_kpi_row(section: dict, data_results: dict, theme: dict) -> str:
    """Render a KPI card row."""
    kpis = section.get("kpis", [])
    cards = []
    for kpi in kpis:
        label = esc(kpi.get("label", ""))
        field = kpi.get("field", "")
        fmt = kpi.get("format", "number")
        target = kpi.get("target")

        value = data_results.get(field)
        display = esc(format_value(value, fmt))

        progress_html = ""
        if value is not None and target is not None:
            target_f = float(target)
            if target_f != 0:
                pct = max(0, min(float(value) / target_f * 100, 100))
                color = "#27ae60" if pct >= 100 else "#f39c12" if pct >= 70 else "#e74c3c"
                progress_html = (
                    f'<div style="margin-top:8px;height:4px;background:{theme["border"]};'
                    f'border-radius:2px;"><div style="width:{pct:.0f}%;height:100%;'
                    f'background:{color};border-radius:2px;"></div></div>'
                )
            progress_html += (
                f'<div style="font-size:11px;color:{theme["text_secondary"]};'
                f'margin-top:4px;">Target: {esc(format_value(target, fmt))}</div>'
            )

        cards.append(
            f'<div style="background:{theme["kpi_bg"]};border:1px solid {theme["border"]};'
            f"box-shadow:{theme['shadow']};border-radius:8px;padding:20px;text-align:center;"
            f'flex:1;min-width:150px;">'
            f'<div style="color:{theme["text_secondary"]};font-size:13px;margin-bottom:6px;">{label}</div>'
            f'<div style="color:{theme["text"]};font-size:26px;font-weight:700;">{display}</div>'
            f"{progress_html}"
            f"</div>"
        )

    return '<div style="display:flex;gap:16px;margin-bottom:24px;flex-wrap:wrap;">' + "\n".join(cards) + "</div>"


def render_chart(section: dict, con: duckdb.DuckDBPyConnection, theme: dict) -> str:
    """Render an ECharts chart section."""
    ds = section.get("data_source", {})
    sql = ds.get("sql", "")
    chart_type = section.get("chart_type", "line")
    x_field = ds.get("x", "")
    y_fields = ds.get("y", [])
    title = esc(section.get("title", ""))
    chart_id = section.get("id", f"chart-{id(section) % 100000}")

    # Execute SQL
    data: list[dict] = []
    if sql:
        try:
            # NOTE: 'sql' is intentionally a raw DuckDB SQL query written by the Agent.
            result = con.execute(sql)
            columns = [desc[0] for desc in result.description]
            rows = result.fetchall()
            data = [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            return f'<div style="color:red;padding:16px;">Chart error: {esc(e)}</div>'

    # Build ECharts option
    echarts_type_map = {
        "line": "line",
        "area": "line",
        "bar": "bar",
        "column": "bar",
        "pie": "pie",
        "donut": "pie",
        "scatter": "scatter",
        "treemap": "treemap",
        "radar": "radar",
        "funnel": "funnel",
    }
    echarts_type = echarts_type_map.get(chart_type, "line")

    option: dict = {}

    if chart_type in ("pie", "donut"):
        agg: dict[str, float] = {}
        for row in data:
            key = str(row.get(x_field, ""))
            val = sum(_safe_float(row.get(yf, 0)) for yf in y_fields)
            agg[key] = agg.get(key, 0) + val
        pie_data = [{"name": k, "value": v} for k, v in agg.items()]
        radius = ["40%", "70%"] if chart_type == "donut" else ["0%", "70%"]
        option = {
            "title": {"text": section.get("title", ""), "textStyle": {"color": theme["text"]}},
            "tooltip": {"trigger": "item"},
            "color": json.loads(theme["palette"]),
            "series": [{"type": "pie", "radius": radius, "data": pie_data}],
        }
    else:
        x_values: list[str] = []
        seen: set[str] = set()
        for row in data:
            xv = str(row.get(x_field, ""))
            if xv not in seen:
                x_values.append(xv)
                seen.add(xv)

        series = []
        for yf in y_fields:
            y_vals = []
            for xv in x_values:
                matched = [r for r in data if str(r.get(x_field, "")) == xv]
                y_vals.append(sum(_safe_float(r.get(yf, 0)) for r in matched) if matched else 0)
            s: dict = {"name": yf, "type": echarts_type, "data": y_vals}
            if chart_type == "area":
                s["areaStyle"] = {}
            series.append(s)

        option = {
            "title": {"text": section.get("title", ""), "textStyle": {"color": theme["text"]}},
            "tooltip": {"trigger": "axis"},
            "color": json.loads(theme["palette"]),
            "xAxis": {"type": "category", "data": x_values, "axisLabel": {"color": theme["text_secondary"]}},
            "yAxis": {"type": "value", "axisLabel": {"color": theme["text_secondary"]}},
            "series": series,
        }

    option_json = json.dumps(option).replace("</script", r"<\/script")

    return (
        f'<div style="background:{theme["card_bg"]};border:1px solid {theme["border"]};'
        f'box-shadow:{theme["shadow"]};border-radius:8px;padding:16px;margin-bottom:24px;">'
        f'<div id="{chart_id}" style="width:100%;height:350px;"></div></div>\n'
        f'<script data-chart-id="{chart_id}" type="application/json">{option_json}</script>'
    )


def render_table(section: dict, con: duckdb.DuckDBPyConnection, theme: dict) -> str:
    """Render a data table section."""
    ds = section.get("data_source", {})
    sql = ds.get("sql", "")
    columns = section.get("columns", [])
    title = esc(section.get("title", ""))
    highlight = section.get("highlight")

    if not sql:
        return ""

    try:
        # NOTE: 'sql' is intentionally a raw DuckDB SQL query written by the Agent.
        result = con.execute(sql)
        all_cols = [desc[0] for desc in result.description]
        rows = result.fetchall()
    except Exception as e:
        return f'<div style="color:red;padding:16px;">Table error: {esc(e)}</div>'

    if not columns:
        columns = all_cols

    header_cells = "".join(f'<th style="padding:10px 12px;text-align:left;border-bottom:2px solid {theme["border"]};color:{theme["text"]};font-weight:600;">{esc(c)}</th>' for c in columns)

    body_rows = ""
    for row_tuple in rows:
        row = dict(zip(all_cols, row_tuple))
        cells = ""
        for c in columns:
            val = row.get(c, "")
            style = f"padding:8px 12px;border-bottom:1px solid {theme['border']};color:{theme['text']};"

            if highlight and c == highlight.get("column"):
                condition = highlight.get("condition", "")
                cell_val = row.get(c)
                try:
                    if condition.startswith("<") and cell_val is not None and float(cell_val) < float(condition[1:]):
                        style += f"color:{highlight.get('color', '#e74c3c')};font-weight:600;"
                    elif condition.startswith(">") and cell_val is not None and float(cell_val) > float(condition[1:]):
                        style += f"color:{highlight.get('color', '#27ae60')};font-weight:600;"
                except (ValueError, TypeError):
                    pass

            cells += f"<td style='{style}'>{esc(val)}</td>"
        body_rows += f"<tr>{cells}</tr>"

    return (
        f'<div style="background:{theme["card_bg"]};border:1px solid {theme["border"]};'
        f"box-shadow:{theme['shadow']};border-radius:8px;padding:16px;margin-bottom:24px;"
        f'overflow-x:auto;">'
        f'<h3 style="color:{theme["text"]};margin:0 0 12px 0;font-size:16px;">{title}</h3>'
        f'<table style="width:100%;border-collapse:collapse;font-size:14px;">'
        f"<thead><tr>{header_cells}</tr></thead>"
        f"<tbody>{body_rows}</tbody>"
        f"</table></div>"
    )


def render_text(section: dict, theme: dict) -> str:
    """Render a text section with basic Markdown support."""
    title = esc(section.get("title", ""))
    content = section.get("content", "")

    # Normalize literal \n to actual newlines, then split on real newlines
    content = content.replace("\\n", "\n")
    lines = content.split("\n")
    html_parts = []
    in_list = False
    for line in lines:
        line = line.strip()
        if not line:
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            html_parts.append("<br>")
        elif line.startswith("**") and line.endswith("**") and len(line) > 4:
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            html_parts.append(f"<strong>{esc(line[2:-2])}</strong>")
        elif line.startswith("- "):
            if not in_list:
                html_parts.append("<ul style='margin:4px 0;padding-left:20px;'>")
                in_list = True
            html_parts.append(f"<li>{_inline_bold(esc(line[2:]))}</li>")
        else:
            if in_list:
                html_parts.append("</ul>")
                in_list = False
            html_parts.append(f"<p style='margin:4px 0;'>{_inline_bold(esc(line))}</p>")
    if in_list:
        html_parts.append("</ul>")

    content_html = "\n".join(html_parts)

    return (
        f'<div style="background:{theme["card_bg"]};border:1px solid {theme["border"]};'
        f'box-shadow:{theme["shadow"]};border-radius:8px;padding:20px;margin-bottom:24px;">'
        f'<h3 style="color:{theme["text"]};margin:0 0 12px 0;font-size:16px;">{title}</h3>'
        f'<div style="color:{theme["text_secondary"]};font-size:14px;line-height:1.6;">'
        f"{content_html}</div></div>"
    )


# ---------------------------------------------------------------------------
# Built-in templates
# ---------------------------------------------------------------------------

BUILTIN_TEMPLATES: dict[str, dict] = {
    "executive-summary": {
        "name": "executive-summary",
        "title": "{{title}}",
        "theme": "corporate",
        "sections": [
            {
                "type": "kpi_row",
                "kpis": [
                    {"label": "Total Revenue", "field": "kpi_revenue", "format": "currency"},
                    {"label": "Total Orders", "field": "kpi_orders", "format": "number"},
                    {"label": "Avg Value", "field": "kpi_avg", "format": "currency"},
                    {"label": "Customers", "field": "kpi_customers", "format": "number"},
                ],
            },
            {
                "type": "chart",
                "chart_type": "line",
                "title": "Revenue Trend",
                "data_source": {
                    "sql": "SELECT order_date as date, SUM(revenue) as revenue FROM data GROUP BY order_date ORDER BY order_date",
                    "x": "date",
                    "y": ["revenue"],
                },
            },
            {
                "type": "text",
                "title": "Key Highlights",
                "content": "{{highlights}}",
            },
        ],
    },
    "weekly-report": {
        "name": "weekly-report",
        "title": "Weekly Report - {{week}}",
        "theme": "light",
        "sections": [
            {
                "type": "kpi_row",
                "kpis": [
                    {"label": "This Week Revenue", "field": "kpi_revenue", "format": "currency"},
                    {"label": "Orders", "field": "kpi_orders", "format": "number"},
                    {"label": "WoW Change", "field": "kpi_wow", "format": "percent"},
                ],
            },
            {
                "type": "chart",
                "chart_type": "area",
                "title": "Daily Revenue",
                "data_source": {
                    "sql": "SELECT order_date as date, SUM(revenue) as revenue FROM data GROUP BY order_date ORDER BY order_date",
                    "x": "date",
                    "y": ["revenue"],
                },
            },
            {
                "type": "table",
                "title": "Performance by Category",
                "columns": ["category", "revenue", "orders"],
                "data_source": {
                    "sql": "SELECT category, SUM(revenue) as revenue, COUNT(*) as orders FROM data GROUP BY category ORDER BY revenue DESC",
                },
            },
            {
                "type": "text",
                "title": "Next Week Plan",
                "content": "{{next_week_plan}}",
            },
        ],
    },
    "kpi-scorecard": {
        "name": "kpi-scorecard",
        "title": "KPI Scorecard - {{period}}",
        "theme": "corporate",
        "sections": [
            {
                "type": "kpi_row",
                "kpis": [
                    {"label": "Revenue", "field": "kpi_revenue", "format": "currency", "target": 100000},
                    {"label": "Orders", "field": "kpi_orders", "format": "number", "target": 500},
                    {"label": "Avg Order", "field": "kpi_avg", "format": "currency", "target": 200},
                    {"label": "Satisfaction", "field": "kpi_satisfaction", "format": "percent", "target": 90},
                ],
            },
            {
                "type": "chart",
                "chart_type": "bar",
                "title": "KPI Achievement",
                "data_source": {
                    "sql": "SELECT category, SUM(revenue) as value FROM data GROUP BY category ORDER BY value DESC LIMIT 5",
                    "x": "category",
                    "y": ["value"],
                },
            },
        ],
    },
}


# ---------------------------------------------------------------------------
# Full report HTML generation
# ---------------------------------------------------------------------------


def generate_report(
    template: dict,
    con: duckdb.DuckDBPyConnection,
    theme_name: str,
    params: dict,
) -> str:
    """Generate the complete report HTML."""
    theme = THEMES.get(theme_name, THEMES["light"])
    title_raw = template.get("title", "Report")

    # Replace placeholders in title
    for key, val in params.items():
        title_raw = title_raw.replace(f"{{{{{key}}}}}", str(val))
    title = esc(title_raw)

    sections = template.get("sections", [])

    # Pre-compute KPI values from data
    data_results = _compute_kpis(con)

    # Render sections
    section_parts: list[str] = []
    for section in sections:
        stype = section.get("type", "")
        if stype == "kpi_row":
            section_parts.append(render_kpi_row(section, data_results, theme))
        elif stype == "chart":
            section_parts.append(render_chart(section, con, theme))
        elif stype == "table":
            section_parts.append(render_table(section, con, theme))
        elif stype == "text":
            # Replace placeholders in text content
            content = section.get("content", "")
            for key, val in params.items():
                content = content.replace(f"{{{{{key}}}}}", str(val))
            section["content"] = content
            section_parts.append(render_text(section, theme))

    sections_html = "\n".join(section_parts)

    return f"""<!DOCTYPE html>
<html lang="auto">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{title}</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    background: {theme["bg"]};
    color: {theme["text"]};
    padding: 40px;
    max-width: 1000px;
    margin: 0 auto;
  }}
  .report-header {{
    margin-bottom: 32px;
    padding-bottom: 20px;
    border-bottom: 3px solid {theme["accent"]};
  }}
  .report-header h1 {{
    font-size: 28px;
    font-weight: 700;
    color: {theme["text"]};
    margin-bottom: 4px;
  }}
  .report-header .subtitle {{
    font-size: 14px;
    color: {theme["text_secondary"]};
  }}
  .report-footer {{
    margin-top: 40px;
    padding-top: 16px;
    border-top: 1px solid {theme["border"]};
    font-size: 12px;
    color: {theme["text_secondary"]};
    text-align: center;
  }}
  @media print {{
    body {{ padding: 20px; }}
    .report-header {{ border-bottom-width: 2px; }}
  }}
</style>
</head>
<body>
<div class="report-header">
  <h1>{title}</h1>
  <div class="subtitle">Generated by DeerFlow Report Template</div>
</div>
{sections_html}
<div class="report-footer">
Report generated automatically &mdash; DeerFlow
</div>
<script src="{ECHARTS_CDN}"></script>
<script>
document.addEventListener('DOMContentLoaded', function() {{
  document.querySelectorAll('script[data-chart-id]').forEach(function(el) {{
    var chartId = el.getAttribute('data-chart-id');
    var dom = document.getElementById(chartId);
    if (!dom) return;
    var chart = echarts.init(dom);
    chart.setOption(JSON.parse(el.textContent));
    window.addEventListener('resize', function() {{ chart.resize(); }});
  }});
}});
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a formatted report from a template definition")
    parser.add_argument(
        "--template",
        required=True,
        help="Path to template JSON file, or a built-in template name (executive-summary, weekly-report, kpi-scorecard)",
    )
    parser.add_argument(
        "--data-files",
        nargs="*",
        default=[],
        help="Data file(s) to load (CSV, Excel, or JSON)",
    )
    parser.add_argument(
        "--output-file",
        required=True,
        help="Path to output HTML file",
    )
    parser.add_argument(
        "--params",
        default="{}",
        help="JSON string of parameter values for template placeholders",
    )
    parser.add_argument(
        "--theme",
        default=None,
        choices=["light", "dark", "corporate"],
        help="Override theme (default: use template value or 'light')",
    )

    args = parser.parse_args()

    # Validate output path
    output_path = validate_path(args.output_file, "output file")

    # Load template
    if args.template in BUILTIN_TEMPLATES:
        template = BUILTIN_TEMPLATES[args.template]
    elif os.path.exists(args.template):
        validate_path(args.template, "template file")
        with open(args.template, encoding="utf-8") as f:
            template = json.load(f)
    else:
        print(f"Error: Template not found: {args.template}")
        print(f"Built-in templates: {', '.join(BUILTIN_TEMPLATES.keys())}")
        sys.exit(1)

    theme_name = args.theme or template.get("theme", "light")

    # Parse params
    try:
        params = json.loads(args.params)
    except json.JSONDecodeError:
        params = {}

    # Load data
    con = duckdb.connect()
    try:
        if args.data_files:
            load_data_files(con, args.data_files)

        # Generate report
        html_output = generate_report(template, con, theme_name, params)
    finally:
        con.close()

    # Write output
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_output)

    section_count = len(template.get("sections", []))
    print(f"Report generated: {output_path}")
    print(f"  Template: {template.get('name', args.template)}")
    print(f"  Theme: {theme_name}")
    print(f"  Sections: {section_count}")


if __name__ == "__main__":
    main()
